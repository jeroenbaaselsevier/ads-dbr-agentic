"""
Build Excel workbook + charts for the country co-affiliation analysis.
Run from repo root with the .venv activated.
"""

import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
from pathlib import Path
import warnings
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

SRC  = Path('tmp/coaffil')
OUT  = Path('output/coaffiliation')
OUT.mkdir(parents=True, exist_ok=True)

# ── Load all CSVs ─────────────────────────────────────────────────────────────
paper_year       = pd.read_csv(SRC / 'paper_trend_by_year.csv')
author_year      = pd.read_csv(SRC / 'author_trend_by_year.csv')
paper_year_field = pd.read_csv(SRC / 'paper_trend_by_year_and_field.csv')
paper_subfield   = pd.read_csv(SRC / 'paper_breakdown_by_subfield.csv')
paper_year_sf    = pd.read_csv(SRC / 'paper_trend_by_year_and_subfield.csv')
author_subfield  = pd.read_csv(SRC / 'author_breakdown_by_subfield.csv')
author_year_sf   = pd.read_csv(SRC / 'author_trend_by_year_and_subfield.csv')
author_cohort    = pd.read_csv(SRC / 'author_first_mca_cohort_by_year_field.csv')
pairs_overall    = pd.read_csv(SRC / 'country_pairs_overall.csv')
pairs_field      = pd.read_csv(SRC / 'country_pairs_by_field.csv')
sample_200       = pd.read_csv(SRC / 'sample_200_papers_multi_country_author.csv', encoding='windows-1252')

# Load new country analysis CSVs if available
try:
    country_summary = pd.read_csv(OUT / 'country_summary_basic.csv')
    country_pairs_null = pd.read_csv(OUT / 'country_pairs_with_null_model.csv')
    country_with_partners = pd.read_csv(OUT / 'country_summary_with_partners.csv')
except:
    country_summary = None
    country_pairs_null = None
    country_with_partners = None

# Filter to 1996-2025
paper_year       = paper_year[paper_year['sort_year'].between(1996, 2025)]
author_year      = author_year[author_year['sort_year'].between(1996, 2025)]
paper_year_field = paper_year_field[paper_year_field['sort_year'].between(1996, 2025)]
paper_year_sf    = paper_year_sf[paper_year_sf['sort_year'].between(1996, 2025)]
author_year_sf   = author_year_sf[author_year_sf['sort_year'].between(1996, 2025)]
author_cohort    = author_cohort[author_cohort['first_mca_year'].between(1996, 2025)]

# Cast numeric types
for df in [paper_year, author_year, paper_year_field, paper_subfield,
           paper_year_sf, author_subfield, author_year_sf, author_cohort,
           pairs_overall, pairs_field]:
    for col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(df[col])
# Sort year-based frames
paper_year       = paper_year.sort_values('sort_year')
author_year      = author_year.sort_values('sort_year')
paper_year_field = paper_year_field.sort_values(['sort_year', 'Field'])
author_year_sf   = author_year_sf.sort_values(['sort_year', 'author_subfield'])
paper_year_sf    = paper_year_sf.sort_values(['sort_year', 'Subfield'])

# ── Clean country-pair overall: drop dirty codes ──────────────────────────────
# Remove rows where country codes have trailing spaces or are clearly junk
pairs_overall = pairs_overall[
    pairs_overall['country1'].str.len() == 3
].copy()
pairs_overall_clean = pairs_overall.groupby(
    ['country1', 'country2'], as_index=False
).agg({'n_author_paper_events': 'sum', 'n_papers': 'sum', 'n_authors': 'sum'})
pairs_overall_clean = pairs_overall_clean.sort_values(
    'n_author_paper_events', ascending=False
).reset_index(drop=True)

# ── Field-level pivot for paper trend ─────────────────────────────────────────
piv_paper_field = paper_year_field.pivot_table(
    index='sort_year', columns='Field',
    values='pct_papers_mca', aggfunc='first'
).sort_index()

# Author year_field (use subfield table aggregated to field level)
author_year_field = (
    author_year_sf
    .groupby(['sort_year', 'author_field'], as_index=False)
    .agg(
        pool_authors_active=('pool_authors_active', 'sum'),
        pool_authors_with_mca=('pool_authors_with_mca', 'sum'),
    )
)
author_year_field['pct_authors_mca'] = (
    100.0 * author_year_field['pool_authors_with_mca']
    / author_year_field['pool_authors_active']
).round(4)
piv_author_field = author_year_field.pivot_table(
    index='sort_year', columns='author_field',
    values='pct_authors_mca', aggfunc='first'
).sort_index()


def to_5y_period_start(year):
    # Use 1996-aligned 5-year bins: 1996-2000, 2001-2005, ...
    return ((int(year) - 1996) // 5) * 5 + 1996


def period_label(start):
    return f"{int(start)}-{int(start)+4}"


# Consistent author heatmap tables used in both figures and Excel
author_year_field_heat = author_year_field.copy()
author_year_field_heat['period_start'] = author_year_field_heat['sort_year'].apply(to_5y_period_start)
author_year_field_heat['period_label'] = author_year_field_heat['period_start'].apply(period_label)

heat_data_author_field_full = (
    author_year_field_heat
    .groupby(['author_field', 'period_start', 'period_label'], as_index=False)
    .apply(lambda g: 100.0 * g['pool_authors_with_mca'].sum() / max(g['pool_authors_active'].sum(), 1))
)
heat_data_author_field_full.columns = ['author_field', 'period_start', 'period_label', 'pct']
heat_data_author_field_full = heat_data_author_field_full.sort_values(['author_field', 'period_start'])

piv_author_field_period = heat_data_author_field_full.pivot(
    index='author_field', columns='period_label', values='pct'
).fillna(0)

author_year_sf_heat = author_year_sf.copy()
author_year_sf_heat['period_start'] = author_year_sf_heat['sort_year'].apply(to_5y_period_start)
author_year_sf_heat['period_label'] = author_year_sf_heat['period_start'].apply(period_label)

heat_data_author_sf_full = (
    author_year_sf_heat
    .groupby(['author_subfield', 'period_start', 'period_label'], as_index=False)
    .apply(lambda g: 100.0 * g['pool_authors_with_mca'].sum() / max(g['pool_authors_active'].sum(), 1))
)
heat_data_author_sf_full.columns = ['author_subfield', 'period_start', 'period_label', 'pct']
heat_data_author_sf_full = heat_data_author_sf_full.sort_values(['author_subfield', 'period_start'])

piv_author_subfield_period = heat_data_author_sf_full.pivot(
    index='author_subfield', columns='period_label', values='pct'
).fillna(0)

# ── CHART 1: Overall paper % trend ───────────────────────────────────────────
fig, ax = plt.subplots(figsize=(10, 5))
ax.plot(paper_year['sort_year'], paper_year['pct_papers_mca'],
        color='steelblue', linewidth=2, marker='o', markersize=3)
ax.set_title('Papers with ≥1 author from 2+ countries (%)', fontsize=13)
ax.set_xlabel('Year')
ax.set_ylabel('% of all papers')
ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.1f%%'))
ax.grid(axis='y', alpha=0.4)
ax.set_xlim(paper_year['sort_year'].min(), paper_year['sort_year'].max())
fig.tight_layout()
fig.savefig(OUT / 'fig1_paper_trend_overall.png', dpi=150)
plt.close()

# ── CHART 2: Overall author % trend ──────────────────────────────────────────
fig, ax = plt.subplots(figsize=(10, 5))
ax.plot(author_year['sort_year'], author_year['pct_authors_mca'],
        color='darkorange', linewidth=2, marker='o', markersize=3)
ax.set_title('Pool authors (5+ papers) with ≥1 multi-country paper in that year (%)', fontsize=12)
ax.set_xlabel('Year')
ax.set_ylabel('% of active pool authors')
ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.1f%%'))
ax.grid(axis='y', alpha=0.4)
ax.set_xlim(author_year['sort_year'].min(), author_year['sort_year'].max())
fig.tight_layout()
fig.savefig(OUT / 'fig2_author_trend_overall.png', dpi=150)
plt.close()

# ── CHART 2b: Paper-level sensitivity (all authors vs 5+ pool-only) ─────────
if 'pct_papers_mca_pool' in paper_year.columns:
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(paper_year['sort_year'], paper_year['pct_papers_mca'],
        color='steelblue', linewidth=2, marker='o', markersize=3,
        label='All authors (current paper indicator)')
    ax.plot(paper_year['sort_year'], paper_year['pct_papers_mca_pool'],
        color='crimson', linewidth=2, marker='s', markersize=3,
        label='5+ pool authors only (sensitivity)')
    ax.set_title('Paper MCA sensitivity: all-authors vs pool-only indicator', fontsize=12)
    ax.set_xlabel('Year')
    ax.set_ylabel('% of papers')
    ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.1f%%'))
    ax.grid(axis='y', alpha=0.4)
    ax.set_xlim(paper_year['sort_year'].min(), paper_year['sort_year'].max())
    ax.legend(fontsize=9, loc='upper left', framealpha=0.9)
    fig.tight_layout()
    fig.savefig(OUT / 'fig2b_paper_sensitivity_all_vs_pool.png', dpi=150)
    plt.close()

# ── CHART 3: Paper % by Field over time ──────────────────────────────────────
fields = sorted(piv_paper_field.columns.tolist())
n = len(fields)
cmap = plt.cm.get_cmap('tab20', n)
colors = [cmap(i) for i in range(n)]

fig, ax = plt.subplots(figsize=(13, 7))
for i, field in enumerate(fields):
    if field in piv_paper_field.columns:
        ax.plot(piv_paper_field.index, piv_paper_field[field],
                label=field, color=colors[i], linewidth=1.5)
ax.set_title('Papers with multi-country author (%) by SM Field', fontsize=13)
ax.set_xlabel('Year')
ax.set_ylabel('% of papers in field')
ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.1f%%'))
ax.grid(axis='y', alpha=0.3)
ax.set_xlim(piv_paper_field.index.min(), piv_paper_field.index.max())
ax.legend(fontsize=7, ncol=2, loc='upper left', framealpha=0.8)
fig.tight_layout()
fig.savefig(OUT / 'fig3_paper_trend_by_field.png', dpi=150)
plt.close()

# ── CHART 4: Author % by Field over time ─────────────────────────────────────
fields_a = sorted(piv_author_field.columns.tolist())
n_a = len(fields_a)
cmap_a = plt.cm.get_cmap('tab20', n_a)
colors_a = [cmap_a(i) for i in range(n_a)]

fig, ax = plt.subplots(figsize=(13, 7))
for i, field in enumerate(fields_a):
    if field in piv_author_field.columns:
        ax.plot(piv_author_field.index, piv_author_field[field],
                label=field, color=colors_a[i], linewidth=1.5)
ax.set_title('Pool authors with multi-country paper (%) by SM Field', fontsize=13)
ax.set_xlabel('Year')
ax.set_ylabel('% of active pool authors in field')
ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.1f%%'))
ax.grid(axis='y', alpha=0.3)
ax.set_xlim(piv_author_field.index.min(), piv_author_field.index.max())
ax.legend(fontsize=7, ncol=2, loc='upper left', framealpha=0.8)
fig.tight_layout()
fig.savefig(OUT / 'fig4_author_trend_by_field.png', dpi=150)
plt.close()

# ── CHART 5: Top 20 country pairs (bar chart) ────────────────────────────────
top20 = pairs_overall_clean.head(20).copy()
top20['pair'] = top20['country1'].str.upper() + ' – ' + top20['country2'].str.upper()
fig, ax = plt.subplots(figsize=(10, 7))
bars = ax.barh(top20['pair'][::-1], top20['n_author_paper_events'][::-1],
               color='steelblue', edgecolor='none')
ax.set_title('Top 20 country pairs by author-paper events (1996–2025)', fontsize=12)
ax.set_xlabel('Author-paper co-occurrence events')
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{int(x):,}'))
ax.grid(axis='x', alpha=0.4)
fig.tight_layout()
fig.savefig(OUT / 'fig5_top20_country_pairs.png', dpi=150)
plt.close()

# ── CHART 6: Heatmap — paper % by Field and decade ───────────────────────────
paper_year_field['decade'] = (paper_year_field['sort_year'] // 5 * 5)
heat_data = (
    paper_year_field
    .groupby(['Field', 'decade'])
    .apply(lambda g: 100.0 * g['papers_with_mca'].sum() / g['total_papers'].sum())
    .reset_index(name='pct')
    .pivot(index='Field', columns='decade', values='pct')
    .fillna(0)
)
heat_data = heat_data.loc[heat_data.mean(axis=1).sort_values(ascending=False).index]

fig, ax = plt.subplots(figsize=(10, 8))
im = ax.imshow(heat_data.values, aspect='auto', cmap='coolwarm')
ax.set_xticks(range(len(heat_data.columns)))
ax.set_xticklabels([str(int(c)) + 's' for c in heat_data.columns], fontsize=9)
ax.set_yticks(range(len(heat_data.index)))
ax.set_yticklabels(heat_data.index, fontsize=9)
for i in range(heat_data.shape[0]):
    for j in range(heat_data.shape[1]):
        v = heat_data.values[i, j]
        ax.text(j, i, f'{v:.1f}%', ha='center', va='center', fontsize=7,
                color='white' if v > heat_data.values.max() * 0.6 else 'black')
plt.colorbar(im, ax=ax, label='% papers with MCA author')
ax.set_title('Multi-country authorship (%) by SM Field and 5-year period', fontsize=12)
fig.tight_layout()
fig.savefig(OUT / 'fig6_field_heatmap.png', dpi=150)
plt.close()


# ── CHART 7: Top-growth fields over time (paper %) ────────────────────────────
# Identify top-5 fields by absolute growth (2025 - 1996)
start_end = piv_paper_field.loc[[piv_paper_field.index.min(), piv_paper_field.index.max()]]
growth = (start_end.iloc[-1] - start_end.iloc[0]).sort_values(ascending=False)
top5_growth_fields = growth.head(5).index.tolist()

fig, ax = plt.subplots(figsize=(12, 6))
# Overall line for context
ax.plot(paper_year['sort_year'], paper_year['pct_papers_mca'],
        color='black', linewidth=2.5, linestyle='--', label='Overall', zorder=10)
colors7 = plt.cm.tab10.colors
for i, field in enumerate(top5_growth_fields):
    vals = piv_paper_field[field]
    ax.plot(vals.index, vals.values,
            color=colors7[i], linewidth=2, marker='o', markersize=3,
            label=f'{field} (+{growth[field]:.1f}pp)')
ax.set_title('Top-5 fastest-growing SM fields: papers with multi-country author (%)', fontsize=12)
ax.set_xlabel('Year')
ax.set_ylabel('% of papers in field')
ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.1f%%'))
ax.grid(axis='y', alpha=0.4)
ax.set_xlim(piv_paper_field.index.min(), piv_paper_field.index.max())
ax.legend(fontsize=9, loc='upper left', framealpha=0.9)
fig.tight_layout()
fig.savefig(OUT / 'fig7_top_growth_fields.png', dpi=150)
plt.close()

# ── CHART 8: 1996 vs 2025 by field (horizontal paired bar) ───────────────────
yrs = piv_paper_field.loc[[1996, 2025]].T.rename(columns={1996: 'y1996', 2025: 'y2025'})
yrs['growth'] = yrs['y2025'] - yrs['y1996']
yrs = yrs.sort_values('y2025')

fig, ax = plt.subplots(figsize=(11, 8))
y_pos = range(len(yrs))
ax.barh(list(y_pos), yrs['y2025'], height=0.6, color='steelblue', label='2025', alpha=0.85)
ax.barh([p + 0.0 for p in y_pos], yrs['y1996'], height=0.35,
        color='lightcoral', label='1996', alpha=0.9)
for i, (_, row) in enumerate(yrs.iterrows()):
    ax.text(row['y2025'] + 0.05, i, f"{row['y2025']:.1f}%", va='center', fontsize=7.5, color='steelblue')
    ax.text(max(row['y1996'] - 0.05, 0.1), i - 0.22, f"{row['y1996']:.1f}%",
            va='center', ha='right', fontsize=7, color='darkred')
ax.set_yticks(list(y_pos))
ax.set_yticklabels(yrs.index, fontsize=9)
ax.set_xlabel('% of papers with ≥1 multi-country author')
ax.set_title('Multi-country authorship rate by SM Field: 1996 vs 2025', fontsize=12)
ax.xaxis.set_major_formatter(mticker.FormatStrFormatter('%.1f%%'))
ax.legend(fontsize=10)
ax.grid(axis='x', alpha=0.4)
fig.tight_layout()
fig.savefig(OUT / 'fig8_field_1996_vs_2025.png', dpi=150)
plt.close()

# ── CHART 9: Author % by Field and period ───────────────────────────────────
heat_data_author_field = piv_author_field_period.copy()
heat_data_author_field = heat_data_author_field.loc[
    heat_data_author_field.mean(axis=1).sort_values(ascending=False).index
]

fig, ax = plt.subplots(figsize=(10, 8))
im = ax.imshow(heat_data_author_field.values, aspect='auto', cmap='coolwarm')
ax.set_xticks(range(len(heat_data_author_field.columns)))
ax.set_xticklabels(list(heat_data_author_field.columns), fontsize=8, rotation=45)
ax.set_yticks(range(len(heat_data_author_field.index)))
ax.set_yticklabels(heat_data_author_field.index, fontsize=9)
for i in range(heat_data_author_field.shape[0]):
    for j in range(heat_data_author_field.shape[1]):
        v = heat_data_author_field.values[i, j]
        ax.text(j, i, f'{v:.1f}%', ha='center', va='center', fontsize=7,
                color='white' if v > heat_data_author_field.values.max() * 0.6 else 'black')
plt.colorbar(im, ax=ax, label='% pool authors with MCA paper')
ax.set_title('Author MCA (%) by SM Field and 5-year period', fontsize=12)
fig.tight_layout()
fig.savefig(OUT / 'fig9_author_field_heatmap.png', dpi=150)
plt.close()

# ── CHART 10: Author % by Subfield and period ────────────────────────────────
heat_data_author_sf = piv_author_subfield_period.copy()
# Keep all subfields for a complete, directly comparable view
heat_data_author_sf = heat_data_author_sf.loc[heat_data_author_sf.mean(axis=1).sort_values(ascending=False).index]

fig_h = max(12, min(30, 0.12 * len(heat_data_author_sf.index)))
fig, ax = plt.subplots(figsize=(10, fig_h))
im = ax.imshow(heat_data_author_sf.values, aspect='auto', cmap='coolwarm')
ax.set_xticks(range(len(heat_data_author_sf.columns)))
ax.set_xticklabels(list(heat_data_author_sf.columns), fontsize=8, rotation=45)
ax.set_yticks(range(len(heat_data_author_sf.index)))
ax.set_yticklabels(heat_data_author_sf.index, fontsize=6)
plt.colorbar(im, ax=ax, label='% pool authors with MCA paper')
ax.set_title('Author MCA (%) by SM Subfield and 5-year period (all subfields)', fontsize=12)
fig.tight_layout()
fig.savefig(OUT / 'fig10_author_subfield_heatmap_top40.png', dpi=150)
plt.close()

print('Charts saved to output/coaffiliation/')

# ── BUILD EXCEL WORKBOOK ──────────────────────────────────────────────────────
xl_path = OUT / 'country_coaffiliation_results.xlsx'

with pd.ExcelWriter(xl_path, engine='openpyxl') as writer:

    def write(df, sheet, index=False, float_fmt='%.4f'):
        df.to_excel(writer, sheet_name=sheet, index=index)
        ws = writer.sheets[sheet]
        # Auto-width columns
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) for c in col_cells if c.value), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)

        # Apply percent display to percentage-like columns (values are already 0-100)
        header_row = 1
        pct_col_idx = []
        for col_idx, cell in enumerate(ws[header_row], start=1):
            header = str(cell.value).lower() if cell.value is not None else ''
            if 'pct' in header or header.startswith('%'):
                pct_col_idx.append(col_idx)

        for col_idx in pct_col_idx:
            for row_idx in range(2, ws.max_row + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                if isinstance(c.value, (int, float)):
                    c.number_format = '0.00\\%'

    def apply_heatmap(ws, first_data_col, last_data_col, first_data_row, last_data_row):
        if last_data_row < first_data_row or last_data_col < first_data_col:
            return
        rng = f"{get_column_letter(first_data_col)}{first_data_row}:{get_column_letter(last_data_col)}{last_data_row}"
        # Blue (low) -> white (mid) -> red (high)
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type='min', start_color='2C7BB6',
                mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                end_type='max', end_color='D7191C'
            )
        )

        # Also show explicit percent symbol in these heatmap cells
        for row_idx in range(first_data_row, last_data_row + 1):
            for col_idx in range(first_data_col, last_data_col + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                if isinstance(c.value, (int, float)):
                    c.number_format = '0.00\\%'

    write(paper_year.sort_values('sort_year'),
          'Paper trend (year)')

    write(author_year.sort_values('sort_year'),
          'Author trend (year)')

    write(piv_paper_field.reset_index(),
          'Paper % by Field+Year', index=False)

    write(piv_author_field.reset_index(),
          'Author % by Field+Year', index=False)

    write(paper_subfield.sort_values('pct_papers_mca', ascending=False),
          'Paper % by Subfield (all yrs)')

    write(author_subfield.sort_values('pct_authors_mca', ascending=False),
          'Author % by Subfield (all yrs)')

    write(paper_year_sf.sort_values(['sort_year', 'Subfield']),
          'Paper % Subfield+Year')

    write(author_year_sf.sort_values(['sort_year', 'author_subfield']),
          'Author % Subfield+Year')

    # Excel heatmaps with identical layout and color scale (all rows, no top-40 cut)
    write(
        piv_author_field_period.reset_index().rename(columns={'author_field': 'Field'}),
        'Heatmap Author Field 5Y',
        index=False,
    )
    ws_field = writer.sheets['Heatmap Author Field 5Y']
    apply_heatmap(
        ws_field,
        first_data_col=2,
        last_data_col=1 + len(piv_author_field_period.columns),
        first_data_row=2,
        last_data_row=1 + len(piv_author_field_period.index),
    )

    write(
        piv_author_subfield_period.reset_index().rename(columns={'author_subfield': 'Subfield'}),
        'Heatmap Author Subfield 5Y',
        index=False,
    )
    ws_sf = writer.sheets['Heatmap Author Subfield 5Y']
    apply_heatmap(
        ws_sf,
        first_data_col=2,
        last_data_col=1 + len(piv_author_subfield_period.columns),
        first_data_row=2,
        last_data_row=1 + len(piv_author_subfield_period.index),
    )

    write(author_cohort.sort_values(['first_mca_year', 'author_field']),
          'Author first MCA cohort')

    write(pairs_overall_clean.head(500),
          'Top country pairs (overall)')

    write(pairs_field.sort_values('n_author_paper_events', ascending=False),
          'Country pairs by Field')

    write(sample_200,
          '200-paper audit sample')

    # Add country analysis sheets if available
    if country_summary is not None:
        write(country_summary.sort_values('authors_with_mca_in_country', ascending=False),
              'Country summary (basic)')
    if country_with_partners is not None:
        write(country_with_partners.sort_values('authors_mca', ascending=False),
              'Country summary (with partners)')
    if country_pairs_null is not None:
        write(country_pairs_null.sort_values('observed_expected_ratio', ascending=False).head(500),
              'Country pairs null model')
        write(country_pairs_null.sort_values('n_author_paper_events', ascending=False).head(500),
              'Country pairs null model vol')

print(f'Excel written to {xl_path}')
